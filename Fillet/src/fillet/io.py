from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Optional, Sequence, Tuple, TypedDict
import csv
import gzip


def open_text(path: str | Path):
    path = str(path)
    return gzip.open(path, "rt", encoding="utf-8", errors="replace") if path.endswith(".gz") else open(path, "r", encoding="utf-8", errors="replace")


@dataclass
class Hit:
    read_id: str
    subject_id: str
    taxid: str
    bitscore: float
    evalue: float
    pident: float
    aln_len: int
    qstart: Optional[int] = None
    qend: Optional[int] = None
    qlen: Optional[int] = None
    qcov: Optional[float] = None
    sstart: Optional[int] = None
    send: Optional[int] = None
    slen: Optional[int] = None
    sample_id: Optional[str] = None
    subject_name: Optional[str] = None
    mismatch: Optional[int] = None
    gapopen: Optional[int] = None
    qseq: Optional[str] = None
    sseq: Optional[str] = None
    btop: Optional[str] = None
    sstrand: Optional[str] = None
    raw: Optional[List[str]] = None
    # Per-alignment uniqueness weight (set by LcaScorer when a uniqueness index is loaded).
    # 1.0 = fully unique k-mers (no downweighting); ~0.001 = conserved locus shared by many taxa.
    uniqueness_weight: float = 1.0

    @property
    def query_coverage(self) -> Optional[float]:
        if self.qcov is not None:
            return self.qcov
        if self.qlen and self.qstart is not None and self.qend is not None and self.qlen > 0:
            return min(1.0, (abs(self.qend - self.qstart) + 1) / self.qlen)
        return None

    @property
    def subject_interval(self) -> Optional[Tuple[int, int]]:
        if self.sstart is None or self.send is None:
            return None
        a, b = int(self.sstart), int(self.send)
        return (min(a, b), max(a, b))


BLAST_STANDARD = [
    "qseqid", "sseqid", "pident", "length", "mismatch", "gapopen", "qstart", "qend",
    "sstart", "send", "evalue", "bitscore"
]


def parse_columns(columns: str | Sequence[str] | None) -> List[str]:
    if columns is None:
        # Backward compatible with the original 14-column alpha layout, while also
        # accepting Fillet v0.2 inspection columns appended after sscinames.
        return [
            "qseqid", "saccver", "pident", "length", "mismatch", "gapopen",
            "qstart", "qend", "sstart", "send", "evalue", "bitscore",
            "staxids", "sscinames", "qlen", "qcovhsp", "slen", "qseq", "sseq",
            "stitle", "btop", "sstrand",
        ]
    if isinstance(columns, str):
        return [c.strip() for c in columns.replace(",", " ").split() if c.strip()]
    return list(columns)


def first_taxid(raw: str | None) -> Optional[str]:
    if raw is None:
        return None
    raw = str(raw).strip()
    if not raw or raw in {"N/A", "NA", "0", "-"}:
        return None
    for sep in [";", ","]:
        if sep in raw:
            raw = raw.split(sep, 1)[0].strip()
    return raw or None


def all_taxids(raw: str | None) -> List[str]:
    """Extract all taxids from a semicolon/comma-separated BLAST staxids string.

    Returns an empty list if the value is missing or invalid.
    Used when a BLAST subject maps to multiple taxa (e.g. shared sequences).
    """
    if raw is None:
        return []
    raw = str(raw).strip()
    if not raw or raw in {"N/A", "NA", "0", "-", ""}:
        return []
    parts = raw.replace(";", ",").split(",")
    return [p.strip() for p in parts if p.strip() and p.strip() not in {"0", "N/A", "NA", "-"}]


def read_fasta_lengths(path: str | Path | None) -> Dict[str, int]:
    return {k: len(v) for k, v in read_fasta_records(path).items()}


def read_fasta_records(path: str | Path | None, max_records: Optional[int] = None) -> Dict[str, str]:
    if not path:
        return {}
    records: Dict[str, str] = {}
    cur: Optional[str] = None
    chunks: List[str] = []
    with open_text(path) as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if cur is not None:
                    records[cur] = "".join(chunks)
                    if max_records is not None and len(records) >= max_records:
                        return records
                cur = line[1:].split()[0]
                chunks = []
            else:
                chunks.append(line)
    if cur is not None and (max_records is None or len(records) < max_records):
        records[cur] = "".join(chunks)
    return records


def read_acc2taxid(path: str | Path | None) -> Dict[str, str]:
    if not path:
        return {}
    mapping: Dict[str, str] = {}
    with open_text(path) as fh:
        first = True
        for line in fh:
            if not line.strip() or line.startswith("#"):
                continue
            parts = line.rstrip("\n").split("\t")
            if len(parts) < 2:
                parts = line.strip().split()
            if len(parts) < 2:
                continue
            if first and any(x.lower() in {"accession", "acc", "taxid", "tax_id"} for x in parts[:2]):
                first = False
                continue
            first = False
            acc, tax = parts[0], parts[1]
            mapping[acc] = tax
            if "." in acc:
                mapping.setdefault(acc.split(".", 1)[0], tax)
    return mapping


def _float(x: str | None, default: float = 0.0) -> float:
    try:
        if x is None or x == "":
            return default
        return float(x)
    except Exception:
        return default


def _int(x: str | None, default: Optional[int] = None) -> Optional[int]:
    try:
        if x is None or x == "":
            return default
        return int(float(x))
    except Exception:
        return default


def iter_blast_hits(
    path: str | Path,
    columns: str | Sequence[str] | None = None,
    acc2taxid: Optional[Dict[str, str]] = None,
    query_lengths: Optional[Dict[str, int]] = None,
    sample_id: Optional[str] = None,
    delimiter: str = "\t",
    expand_multi_taxids: bool = True,
) -> Iterator[Hit]:
    """Iterate over BLAST/LAST tabular alignment hits.

    Parameters
    ----------
    expand_multi_taxids:
        When True (default), a single alignment row that carries multiple
        semicolon-separated staxids (e.g. from a shared/multi-species sequence)
        yields one Hit per taxid.  This ensures RELIC-LCA considers all valid
        taxonomic paths rather than discarding them by taking only the first taxid.
    """
    cols = parse_columns(columns)
    idx = {c: i for i, c in enumerate(cols)}
    acc2taxid = acc2taxid or {}
    query_lengths = query_lengths or {}

    def val(parts: List[str], *names: str) -> Optional[str]:
        for name in names:
            i = idx.get(name)
            if i is not None and i < len(parts):
                return parts[i]
        return None

    with open_text(path) as fh:
        for line in fh:
            line = line.rstrip("\n")
            if not line or line.startswith("#"):
                continue
            parts = line.split(delimiter)
            if len(parts) < min(12, len(cols)):
                continue
            read_id = val(parts, "qseqid", "query", "read_id") or parts[0]
            subject_id = val(parts, "saccver", "sseqid", "subject", "acc") or (parts[1] if len(parts) > 1 else "")
            raw_tax = val(parts, "staxids", "taxid", "tax_id")
            # Resolve taxid(s) — prefer staxids column, fall back to acc2taxid lookup
            fallback_taxid = acc2taxid.get(subject_id) or acc2taxid.get(subject_id.split(".", 1)[0])
            if expand_multi_taxids:
                taxid_list = all_taxids(raw_tax)
                if not taxid_list and fallback_taxid:
                    taxid_list = [fallback_taxid]
            else:
                tid = first_taxid(raw_tax) or fallback_taxid
                taxid_list = [tid] if tid else []
            if not taxid_list:
                continue
            pident = _float(val(parts, "pident"), 0.0)
            aln_len = _int(val(parts, "length", "aln_len"), 0) or 0
            mismatch = _int(val(parts, "mismatch"), None)
            gapopen = _int(val(parts, "gapopen"), None)
            qstart = _int(val(parts, "qstart"), None)
            qend = _int(val(parts, "qend"), None)
            sstart = _int(val(parts, "sstart"), None)
            send = _int(val(parts, "send"), None)
            slen = _int(val(parts, "slen", "subject_len"), None)
            qlen = _int(val(parts, "qlen", "query_len"), None) or query_lengths.get(read_id)
            qcov_raw = val(parts, "qcovhsp", "qcov", "qcovs")
            qcov = None
            if qcov_raw is not None:
                qcov = _float(qcov_raw, 0.0)
                if qcov > 1.0:
                    qcov = qcov / 100.0
            elif qlen and qstart is not None and qend is not None:
                qcov = min(1.0, (abs(qend - qstart) + 1) / max(1, qlen))
            evalue = _float(val(parts, "evalue"), 999.0)
            bitscore = _float(val(parts, "bitscore", "score"), 0.0)
            common_kwargs = dict(
                read_id=read_id,
                subject_id=subject_id,
                bitscore=bitscore,
                evalue=evalue,
                pident=pident,
                aln_len=aln_len,
                qstart=qstart,
                qend=qend,
                qlen=qlen,
                qcov=qcov,
                sstart=sstart,
                send=send,
                slen=slen,
                sample_id=sample_id,
                subject_name=val(parts, "stitle", "subject_title", "sscinames", "subject_name", "name"),
                mismatch=mismatch,
                gapopen=gapopen,
                qseq=val(parts, "qseq"),
                sseq=val(parts, "sseq"),
                btop=val(parts, "btop", "BTOP"),
                sstrand=val(parts, "sstrand", "strand"),
                raw=parts,
            )
            for taxid in taxid_list:
                yield Hit(taxid=str(taxid), **common_kwargs)


def _parse_taxid_from_last_name(name: str) -> Optional[str]:
    """Extract taxid from a LAST subject name.

    Handles the ``>accession|taxid:12345`` header format written by
    :func:`fillet.builddb.extract_sequences_blastdbcmd`.  Falls back to
    ``None`` if no taxid token is present.
    """
    if "|taxid:" in name:
        try:
            return name.split("|taxid:", 1)[1].split("|")[0].strip()
        except Exception:
            pass
    return None


def _pident_from_alignment(qseq: str, sseq: str) -> Tuple[float, int, int]:
    """Compute percent identity, alignment length and mismatch count.

    Parameters
    ----------
    qseq, sseq:
        Aligned query and subject strings (same length, gaps as ``'-'``).

    Returns
    -------
    (pident, aln_len, mismatches)
        *pident* = matches / aln_len × 100 (range 0–100).
        *aln_len* = total alignment columns including gap columns.
        *mismatches* = non-gap positions where characters differ.
    """
    assert len(qseq) == len(sseq), "Alignment strings must have equal length"
    aln_len = len(qseq)
    matches = 0
    mismatches = 0
    for q, s in zip(qseq, sseq):
        if q == "-" or s == "-":
            continue
        if q.upper() == s.upper():
            matches += 1
        else:
            mismatches += 1
    pident = matches / max(1, aln_len) * 100.0
    return pident, aln_len, mismatches


def _parse_maf_s_line(line: str):
    """Parse a MAF ``s`` line into its seven fields.

    MAF ``s`` line format::

        s  name  start  size  strand  srcSize  text

    All fields are whitespace-separated.  Returns a tuple::

        (name, start, size, strand, src_size, text)

    where *start*, *size*, and *src_size* are integers.

    Raises
    ------
    ValueError
        If the line does not have the expected number of fields.
    """
    parts = line.split()
    if len(parts) != 7 or parts[0] != "s":
        raise ValueError(f"Malformed MAF s-line: {line!r}")
    name = parts[1]
    start = int(parts[2])
    size = int(parts[3])
    strand = parts[4]
    src_size = int(parts[5])
    text = parts[6]
    return name, start, size, strand, src_size, text


def _maf_coords_to_blast(start: int, size: int, strand: str, src_size: int):
    """Convert 0-based LAST/MAF half-open coords to 1-based BLAST closed coords.

    LAST stores coordinates from the 5′ end of the *plus* strand regardless of
    the hit strand.  For a minus-strand alignment the stored *start* is the
    distance from the plus-strand 5′ end; the actual minus-strand position is
    computed by mirroring against the sequence length.

    Parameters
    ----------
    start:
        0-based start from the MAF ``s`` line.
    size:
        Aligned length (excluding gap characters) from the MAF ``s`` line.
    strand:
        ``'+'`` or ``'-'``.
    src_size:
        Full sequence length from the MAF ``s`` line.

    Returns
    -------
    (start_1b, end_1b)
        1-based closed interval, always with start_1b ≤ end_1b.
    """
    if strand == "+":
        start_1b = start + 1
        end_1b = start + size
    else:
        # LAST convention: for minus strand, start is from the + strand.
        # The actual plus-strand interval is [start, start+size).
        # Converted to minus-strand 1-based: flip around sequence length.
        end_1b = src_size - start          # mirror of 0-based start → 1-based end
        start_1b = src_size - start - size + 1  # mirror of 0-based end → 1-based start
    return start_1b, end_1b


def iter_last_maf_hits(
    path: str | Path,
    *,
    acc2taxid: Optional[Dict[str, str]] = None,
    query_lengths: Optional[Dict[str, int]] = None,
    sample_id: Optional[str] = None,
    max_hits_per_read: int = 0,
    min_score: float = 0.0,
    last_score_to_bitscore_scale: float = 1.0,
    expand_multi_taxids: bool = True,
) -> Iterator[Hit]:
    """Parse LAST MAF-format output into :class:`Hit` objects.

    Reads the output of ``lastal -f MAF …`` and emits one :class:`Hit` per
    alignment block (or multiple hits when a subject name resolves to several
    taxids via semicolons and *expand_multi_taxids* is True).

    MAF block structure::

        a score=NNN EG2=X E=X
        s subj_name   sstart  saln_len  sstrand  sseqlen  salignment
        s query_name  qstart  qaln_len  qstrand  qseqlen  qalignment

    LAST always writes the database (reference) sequence first and the query
    (read) sequence second — the reverse of what one might expect.

    Alignment blocks are separated by blank lines.  Lines beginning with ``#``
    are comment/header lines and are skipped.

    Coordinate conversion
    ---------------------
    LAST uses 0-based half-open intervals measured from the plus-strand 5′ end.
    This function converts them to 1-based closed intervals (BLAST convention).
    For minus-strand subject hits the start/end are flipped so that
    ``sstart ≤ send`` always holds (matching BLAST's ``sstrand`` behaviour).

    Taxid resolution
    ----------------
    The function first tries to extract a taxid from the subject name using
    the ``|taxid:XXXX`` token written by
    :func:`fillet.builddb.extract_sequences_blastdbcmd`.  If that fails it
    falls back to the *acc2taxid* dict keyed by the bare accession (without
    version suffix).  Alignments without any resolvable taxid are skipped.

    Parameters
    ----------
    path:
        Path to a LAST MAF output file.  May be gzip-compressed (``.gz``).
    acc2taxid:
        Optional ``{accession: taxid}`` lookup used as a fallback when the
        subject name does not contain a ``|taxid:`` token.
    query_lengths:
        Optional ``{read_id: length}`` dict for computing ``qcov`` when the
        query length is not available from the MAF block.
    sample_id:
        Value to populate :attr:`Hit.sample_id` on every emitted hit.
    max_hits_per_read:
        Maximum number of hits to emit per read (0 = no limit).  Hits are
        counted in file order; the limit is applied after taxid expansion.
    min_score:
        Minimum raw LAST score (before bitscore scaling).  Blocks with a
        lower score are skipped entirely.
    last_score_to_bitscore_scale:
        Multiply the raw LAST score by this factor to obtain an approximate
        bitscore stored in :attr:`Hit.bitscore`.  Default 1.0 (raw score).
    expand_multi_taxids:
        When True, a subject taxid string containing multiple semicolon-
        separated values yields one :class:`Hit` per taxid (mirroring the
        behaviour of :func:`iter_blast_hits`).

    Yields
    ------
    Hit
        One per alignment block (or per taxid if *expand_multi_taxids*).
    """
    acc2taxid = acc2taxid or {}
    query_lengths = query_lengths or {}
    read_hit_counts: Dict[str, int] = {}

    # --- state machine variables ---
    a_line: Optional[str] = None   # the 'a score=...' line for current block
    s_lines: List[str] = []        # accumulated 's' lines for current block

    def _process_block(a_ln: str, s_lns: List[str]) -> Iterator[Hit]:
        """Yield Hit(s) from a completed MAF alignment block."""
        if len(s_lns) < 2:
            return  # incomplete block (e.g. unmapped query)

        # Parse 'a' line: score and optional E-value
        score = 0.0
        evalue = 999.0
        for token in a_ln.split():
            if token.startswith("score="):
                try:
                    score = float(token[6:])
                except ValueError:
                    pass
            elif token.startswith("E="):
                try:
                    evalue = float(token[2:])
                except ValueError:
                    pass
            elif token.startswith("EG2="):
                pass  # ignore EG2 (expected per sq-giga); use E= for evalue

        if score < min_score:
            return

        # LAST outputs reference (subject/DB) first, query second.
        try:
            s_name, s_start, s_size, s_strand, s_src_size, s_text = _parse_maf_s_line(s_lns[0])
            q_name, q_start, q_size, q_strand, q_src_size, q_text = _parse_maf_s_line(s_lns[1])
        except (ValueError, IndexError):
            return  # malformed block — skip silently

        # --- taxid resolution ---
        raw_taxid = _parse_taxid_from_last_name(s_name)
        if raw_taxid is None:
            # Try acc2taxid fallback using bare accession
            bare_acc = s_name.split("|")[0].split(".")[0]
            raw_taxid = acc2taxid.get(s_name) or acc2taxid.get(bare_acc)

        if expand_multi_taxids:
            taxid_list = all_taxids(raw_taxid)
            if not taxid_list:
                return
        else:
            tid = first_taxid(raw_taxid)
            if not tid:
                return
            taxid_list = [tid]

        # --- coordinate conversion ---
        q_start_1b, q_end_1b = _maf_coords_to_blast(q_start, q_size, q_strand, q_src_size)
        s_start_1b, s_end_1b = _maf_coords_to_blast(s_start, s_size, s_strand, s_src_size)

        # --- alignment metrics ---
        pident, aln_len, mismatches = _pident_from_alignment(q_text, s_text)
        bitscore = score * last_score_to_bitscore_scale

        # --- query coverage ---
        qlen = q_src_size or query_lengths.get(q_name)
        qcov: Optional[float] = None
        if qlen and qlen > 0:
            qcov = min(1.0, q_size / qlen)

        common_kwargs: Dict = dict(
            read_id=q_name,
            subject_id=s_name,
            bitscore=bitscore,
            evalue=evalue,
            pident=pident,
            aln_len=aln_len,
            qstart=q_start_1b,
            qend=q_end_1b,
            qlen=qlen,
            qcov=qcov,
            sstart=s_start_1b,
            send=s_end_1b,
            slen=s_src_size,
            sample_id=sample_id,
            subject_name=s_name,
            mismatch=mismatches,
            gapopen=None,    # MAF doesn't report gap-open count directly
            qseq=q_text,
            sseq=s_text,
            btop=None,
            sstrand=s_strand,
            raw=None,
        )

        for taxid in taxid_list:
            # max_hits_per_read guard
            if max_hits_per_read > 0:
                count = read_hit_counts.get(q_name, 0)
                if count >= max_hits_per_read:
                    return
                read_hit_counts[q_name] = count + 1
            yield Hit(taxid=str(taxid), **common_kwargs)

    # --- main parsing loop ---
    with open_text(path) as fh:
        for raw_line in fh:
            line = raw_line.rstrip("\n")

            if line.startswith("#"):
                # Comment / header line — flush current block if any
                if a_line is not None and s_lines:
                    yield from _process_block(a_line, s_lines)
                    a_line = None
                    s_lines = []
                continue

            if line.startswith("a ") or line == "a":
                # New alignment block — flush the previous one first
                if a_line is not None and s_lines:
                    yield from _process_block(a_line, s_lines)
                a_line = line
                s_lines = []

            elif line.startswith("s "):
                if a_line is not None:
                    s_lines.append(line)

            elif not line.strip():
                # Blank line — flush current block
                if a_line is not None and s_lines:
                    yield from _process_block(a_line, s_lines)
                    a_line = None
                    s_lines = []
            # Other line types (q, p, i, c for quality/probability/info/counts)
            # are silently ignored.

    # Flush the final block (file may not end with a blank line)
    if a_line is not None and s_lines:
        yield from _process_block(a_line, s_lines)


def group_hits_by_read(hits: Iterable[Hit]) -> Iterator[Tuple[str, List[Hit]]]:
    cur_id: Optional[str] = None
    buf: List[Hit] = []
    for h in hits:
        if cur_id is None:
            cur_id = h.read_id
            buf = [h]
        elif h.read_id == cur_id:
            buf.append(h)
        else:
            yield cur_id, buf
            cur_id = h.read_id
            buf = [h]
    if cur_id is not None:
        yield cur_id, buf


def _read_table(path: str | Path) -> List[Dict[str, str]]:
    path = Path(path)
    rows: List[Dict[str, str]] = []
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as e:
            raise RuntimeError("Install Fillet with [xlsx] or install openpyxl to read Excel sample sheets") from e
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))[0:]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append({header[i]: ("" if v is None else str(v)) for i, v in enumerate(r) if i < len(header) and header[i]})
    else:
        with open(path, "r", encoding="utf-8", newline="") as fh:
            sample = fh.read(4096)
            fh.seek(0)
            first = sample.splitlines()[0] if sample.splitlines() else ""
            delim = "\t" if "\t" in first else ","
            reader = csv.DictReader(fh, delimiter=delim)
            rows = [dict(r) for r in reader]
    return rows


def read_sample_sheet(path: str | Path | None) -> Dict[str, Dict[str, str]]:
    """Read CSV/TSV/XLSX sample metadata.

    Expected but flexible columns: sample_id, role/type/control_type, group.
    role values can include sample, negative, positive, environmental.
    """
    if not path:
        return {}
    out: Dict[str, Dict[str, str]] = {}
    for row in _read_table(path):
        sid = row.get("sample_id") or row.get("sample") or row.get("library") or row.get("library_id") or row.get("id")
        if sid:
            out[str(sid)] = {str(k): str(v) for k, v in row.items() if k is not None}
    return out


def read_regional_taxa(path: str | Path | None) -> Dict[str, Dict[str, str]]:
    """Read optional regional/ecological priors TSV/CSV/XLSX.

    Columns: taxid or scientific_name/name, status/presence, habitat, weight, notes.
    Keys are duplicated by taxid and scientific name so either can match.
    """
    if not path:
        return {}
    out: Dict[str, Dict[str, str]] = {}
    for row in _read_table(path):
        norm = {str(k): ("" if v is None else str(v)) for k, v in row.items() if k is not None}
        keys = [norm.get("taxid"), norm.get("tax_id"), norm.get("scientific_name"), norm.get("name"), norm.get("taxon")]
        for k in keys:
            if k:
                out[str(k)] = norm
    return out


def read_support_table(path: str | Path | None) -> "set[str]":
    """Read a taxid/name support table (palynology, fossil, etc.) returning a set of identifiers.

    Accepts TSV/CSV/XLSX with a ``taxid``, ``tax_id``, ``name``, or ``scientific_name`` column,
    or a plain text file with one taxid/name per line.
    """
    return read_contaminants(path)


def read_contaminants(path: str | Path | None) -> "set[str]":
    """Read a contaminants list and return a set of taxids.

    Accepts:
    - Plain text: one taxid or name per line, # comments ignored
    - TSV/CSV/XLSX: column named ``taxid``, ``tax_id``, ``name``, or ``scientific_name``
    """
    if not path:
        return set()
    taxids: set[str] = set()
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xls", ".csv", ".tsv", ".tab"):
        for row in _read_table(path):
            norm = {str(k).lower(): ("" if v is None else str(v).strip()) for k, v in row.items() if k is not None}
            tid = norm.get("taxid") or norm.get("tax_id") or norm.get("name") or norm.get("scientific_name")
            if tid:
                taxids.add(tid)
    else:
        with open(path, "r", encoding="utf-8") as fh:
            for line in fh:
                t = line.strip()
                if t and not t.startswith("#"):
                    taxids.add(t)
    return taxids


class FossilRecord(TypedDict, total=False):
    """One row from a structured fossil occurrence table."""
    taxon: str        # name OR taxid (required)
    site_name: str    # site filter; empty/"*" = all sites
    age_min_bp: Optional[float]  # older bound (higher BP); None = no limit
    age_max_bp: Optional[float]  # younger bound (lower BP); None = no limit
    evidence_type: str  # e.g. "bone", "macrofossil", "archaeozoological"
    source: str       # citation string
    notes: str


def read_fossil_table_structured(path: "str | Path | None") -> "list[FossilRecord]":
    """Read a fossil occurrence table, returning structured FossilRecord dicts.

    Accepts TSV/CSV/XLSX with columns: taxon (required), site_name, age_min_bp,
    age_max_bp, evidence_type, source, notes.  Also accepts a plain text file
    (one taxid/name per line) for backward compatibility — those records have
    no site or age restriction and match every sample.
    """
    if not path:
        return []
    p = Path(path)
    records: list[FossilRecord] = []

    def _float_or_none(s: str) -> Optional[float]:
        try:
            return float(s) if s.strip() else None
        except ValueError:
            return None

    if p.suffix.lower() in (".xlsx", ".xls", ".csv", ".tsv", ".tab"):
        for row in _read_table(path):
            norm = {str(k).lower().strip(): ("" if v is None else str(v).strip())
                    for k, v in row.items() if k is not None}
            taxon = (norm.get("taxon") or norm.get("taxid") or norm.get("tax_id")
                     or norm.get("name") or norm.get("scientific_name", ""))
            if not taxon or str(taxon).startswith("#"):
                continue
            records.append(FossilRecord(
                taxon=taxon,
                site_name=norm.get("site_name", ""),
                age_min_bp=_float_or_none(norm.get("age_min_bp", "")),
                age_max_bp=_float_or_none(norm.get("age_max_bp", "")),
                evidence_type=norm.get("evidence_type", ""),
                source=norm.get("source", ""),
                notes=norm.get("notes", ""),
            ))
    else:
        with open(path, "r", encoding="utf-8") as fh:
            for line in fh:
                t = line.strip()
                if t and not t.startswith("#"):
                    records.append(FossilRecord(
                        taxon=t, site_name="", age_min_bp=None,
                        age_max_bp=None, evidence_type="", source="", notes="",
                    ))
    return records


def filter_fossil_taxa_for_sample(
    records: "list[FossilRecord]",
    site_name: "str | None",
    age_bp: "str | float | None",
) -> "tuple[set[str], list[str]]":
    """Filter fossil records for a specific sample and return matching taxa.

    Parameters
    ----------
    records:
        Records loaded by :func:`read_fossil_table_structured`.
    site_name:
        The sample's ``site_name`` value from the batch metadata.
    age_bp:
        The sample's ``age_BP`` value (years before present; higher = older).
        May be a string (from TSV) or numeric.  ``None`` = no age filtering.

    Returns
    -------
    taxa : set[str]
        Taxon names/taxids whose records match this sample's site and age.
    evidence_texts : list[str]
        Human-readable evidence string per matching record for GUI display.
    """
    site = (site_name or "").strip()
    try:
        age: Optional[float] = float(age_bp) if age_bp not in (None, "") else None
    except (ValueError, TypeError):
        age = None

    taxa: set[str] = set()
    evidence: list[str] = []

    for rec in records:
        rec_site = (rec.get("site_name") or "").strip()
        # Site: empty or "*" matches everything; otherwise exact match required
        if rec_site and rec_site != "*" and rec_site != site:
            continue
        # Age: only filter when the sample has a known age
        if age is not None:
            min_bp = rec.get("age_min_bp")
            max_bp = rec.get("age_max_bp")
            # age_min_bp = older bound (higher BP); age_max_bp = younger bound (lower BP)
            # Record applies when: age_max_bp <= sample_age <= age_min_bp
            if min_bp is not None and age > min_bp:
                continue   # sample is older than the record's oldest bound
            if max_bp is not None and age < max_bp:
                continue   # sample is younger than the record's youngest bound
        taxon = rec.get("taxon", "")
        if not taxon:
            continue
        taxa.add(taxon)
        ev_type = rec.get("evidence_type") or "fossil"
        source = rec.get("source", "")
        text = f"{ev_type}: {taxon}"
        if source:
            text += f" ({source})"
        evidence.append(text)

    return taxa, evidence


def write_tsv(path: str | Path, rows: Iterable[Dict[str, object]], fieldnames: Sequence[str]) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, delimiter="\t", fieldnames=list(fieldnames), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
