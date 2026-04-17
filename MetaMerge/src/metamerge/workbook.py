"""Excel workbook writer for MetaMerge merged results.

Produces a styled, publication-friendly workbook with:
  - README sheet explaining the package and thresholds.
  - Merged_all_taxa sheet: one row per taxon.
  - Classification_logic sheet: threshold definitions used for this run.
  - Library_metadata sheet: the library linker table.
  - Run_summary sheet: compact counts and warnings.
  - Warnings sheet: unmatched libraries/taxa and QC notes.
  - Report_files sheet: paths to long-format plotting tables (optional).

Style notes
-----------
Excel Table objects are intentionally NOT used.  Table objects can corrupt
the XML in some openpyxl versions.  Auto-filter is applied directly to the
data range instead, which achieves the same user-facing filtering behaviour
without the corruption risk.

Column widths are estimated from cell content length with a cap of 40
characters to prevent unwieldy columns.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL    = PatternFill("solid", fgColor="1F4E78")   # Dark blue
HEADER_FONT    = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")   # Light blue
BOLD           = Font(bold=True)


def _autosize_columns(ws) -> None:
    """Set column widths based on maximum cell content length (capped at 40)."""
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        length = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in column_cells
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 10), 40)


def _add_dataframe_sheet(
    wb: Workbook,
    name: str,
    df: pd.DataFrame,
    freeze_cell: str = "A2",
) -> None:
    """Append a sheet with a pandas DataFrame.

    Uses plain auto-filter (not Excel Table objects) to avoid XML corruption.

    Args:
        wb: The openpyxl Workbook to append to.
        name: Sheet name (truncated to 31 characters, the Excel maximum).
        df: DataFrame to write.
        freeze_cell: Cell reference to freeze panes at (default = row 2).
    """
    ws = wb.create_sheet(title=name[:31])
    ws.freeze_panes = freeze_cell

    for c_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=None if pd.isna(value) else value)

    ws.auto_filter.ref = ws.dimensions
    _autosize_columns(ws)


def _add_readme_sheet(wb: Workbook, run_info: dict, config: dict) -> None:
    """Create the README sheet as the first (active) sheet."""
    ws = wb.active
    ws.title = "README"

    lines = [
        ("MetaMerge output workbook", True),
        ("", False),
        ("Purpose", True),
        (
            "This workbook was produced by MetaMerge, which merges a conservative "
            "MEGAN BLASTn count matrix with Holi/metaDMG damage outputs and assigns "
            "aDNA-support categories based on damage signals, read counts, and lineage "
            "consistency.  Additional lines of evidence (ecological plausibility, "
            "biogeographic range, macrofossil data) can be applied as a separate "
            "interpretive layer on top of these DNA-based categories.",
            False,
        ),
        ("", False),
        ("Output sheets", True),
        ("Key_results — most important per-taxon columns: taxon name, common name, "
         "aDNA support status, Holi/metaDMG damage values, MEGAN counts, and per-library "
         "aDNA support.  Optimised for immediate review.", False),
        ("Full_results — all output columns including lineage support, blank detail, "
         "extended Holi fields, and library path metadata.", False),
        ("Classification_logic — threshold definitions used for this run.", False),
        ("Library_metadata — the library linker table supplied by the user.", False),
        ("Run_summary — compact run-level counts and unmatched-taxon warnings.", False),
        ("Warnings — unmatched libraries/taxa and QC notes.", False),
        ("", False),
        ("DNA support categories (most to least supported)", True),
        ("Very high confidence — exact Holi/metaDMG damage support (damage > damage_min, "
         "significance > significance_min, N_reads >= high_confidence_n_reads_min) + strong "
         "MEGAN count support + no strong QC caution.", False),
        ("High confidence — exact Holi/metaDMG damage support + strong MEGAN count "
         "support, without the extra N_reads criterion and/or with mild QC caution.", False),
        ("Supported — exact damage support and/or strong count support and/or "
         "low-rank lineage-consistent support, but below the confidence thresholds.", False),
        ("Tentative — weak or incomplete DNA support that exceeds tentative_min_reads "
         "and is not blank-associated.", False),
        ("Weak support — minimal non-control evidence (>= weak_support_min_reads).", False),
        ("Blank-associated — substantial blank overlap and weak real-library support.", False),
        ("", False),
        ("Thresholds used in this run", True),
    ]
    for key, value in sorted(config["thresholds"].items()):
        lines.append((f"  {key}: {value}", False))

    lines += [
        ("", False),
        ("Run summary", True),
        (f"  Taxa processed: {run_info['n_taxa']}", False),
        (f"  Status counts: {run_info['status_counts']}", False),
    ]

    row = 1
    for text, is_header in lines:
        cell = ws.cell(row=row, column=1, value=text)
        if is_header:
            cell.font = BOLD
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws.column_dimensions["A"].width = 120


def write_workbook(
    merged_df: pd.DataFrame,
    metadata_df: pd.DataFrame,
    run_summary: dict,
    warnings_df: pd.DataFrame,
    outpath: Path,
    config: dict,
    plot_inputs: list[Path] | None = None,
) -> None:
    """Write the main publication-style MetaMerge workbook.

    Args:
        merged_df: The merged taxon DataFrame from ``build_merge``.
        metadata_df: The library linker DataFrame.
        run_summary: Summary dict from ``build_merge``.
        warnings_df: Warnings DataFrame from ``build_warnings_df``.
        outpath: Destination ``.xlsx`` path.
        config: Full MetaMerge config dict.
        plot_inputs: Optional list of plot-input TSV Paths to list in the
            Report_files sheet.
    """
    wb = Workbook()
    _add_readme_sheet(wb, run_summary, config)

    # Classification logic sheet — one row per configurable threshold.
    logic = pd.DataFrame(
        [
            ("damage_min",                        config["thresholds"]["damage_min"],
             "Minimum Holi/metaDMG damage for exact damage support."),
            ("significance_min",                  config["thresholds"]["significance_min"],
             "Minimum Holi/metaDMG significance for exact damage support."),
            ("high_confidence_n_reads_min",       config["thresholds"]["high_confidence_n_reads_min"],
             "Minimum Holi/metaDMG N_reads for Very high confidence."),
            ("strong_count_min_reads",            config["thresholds"]["strong_count_min_reads"],
             "Minimum MEGAN count in at least one non-control library for strong count support."),
            ("strong_count_min_libraries",        config["thresholds"]["strong_count_min_libraries"],
             "Minimum number of non-control libraries positive for strong count support."),
            ("blank_absolute_min",                config["thresholds"]["blank_absolute_min"],
             "Absolute blank-count threshold for blank concern."),
            ("blank_relative_min",                config["thresholds"]["blank_relative_min"],
             "Relative blank/real-count threshold for blank concern."),
            ("qc_alignments_per_read_clean_max",  config["thresholds"]["qc_alignments_per_read_clean_max"],
             "Maximum N_alignments/N_reads considered clean."),
            ("qc_alignments_per_read_caution_max", config["thresholds"]["qc_alignments_per_read_caution_max"],
             "Above this multi-mapping ratio, treated as strong caution."),
            ("qc_abs_rho_clean_max",              config["thresholds"]["qc_abs_rho_clean_max"],
             "Maximum |rho_Ac| considered clean."),
            ("qc_abs_rho_caution_max",            config["thresholds"]["qc_abs_rho_caution_max"],
             "Above this |rho_Ac|, treated as strong caution."),
            ("lineage_max_steps",                 config["thresholds"]["lineage_max_steps"],
             "Maximum rank steps between focal and lineage-support taxon."),
            ("lineage_max_rank_level",            config["thresholds"]["lineage_max_rank_level"],
             "Broadest rank allowed for lineage support."),
        ],
        columns=["parameter", "value", "description"],
    )

    run_summary_df = pd.DataFrame(
        [{"metric": "n_taxa", "value": run_summary["n_taxa"]}]
        + [{"metric": f"status__{k}", "value": v}
           for k, v in run_summary["status_counts"].items()]
        + [{"metric": "unmatched_taxa_examples",
            "value": "; ".join(run_summary.get("unmatched_taxa_examples", []))}]
    )

    # ── Key_results sheet: most important columns for immediate review ────────
    _KEY_COLS_PRIORITY = [
        "scientific_name", "common_name", "tax_rank", "broad_group",
        "aDNA_support_status", "support_basis_summary",
        "Holi_best_damage", "Holi_best_significance", "Holi_best_N_reads",
        "Holi_best_library",
        "Holi_exact_damage_sig_libraries_n", "Holi_exact_damage_sig_libraries",
        "megan_max_count", "megan_positive_libraries_n",
        "blank_ratio",
    ]
    count_cols       = [c for c in merged_df.columns if c.startswith("count__")]
    support_lib_cols = [c for c in merged_df.columns if c.startswith("aDNA_support_lib__")]
    key_cols = (
        [c for c in _KEY_COLS_PRIORITY if c in merged_df.columns]
        + count_cols
        + support_lib_cols
    )
    _add_dataframe_sheet(wb, "Key_results",  merged_df[key_cols])
    _add_dataframe_sheet(wb, "Full_results", merged_df)
    _add_dataframe_sheet(wb, "Classification_logic", logic)
    _add_dataframe_sheet(wb, "Library_metadata",   metadata_df)
    _add_dataframe_sheet(wb, "Run_summary",        run_summary_df)
    _add_dataframe_sheet(
        wb, "Warnings",
        warnings_df if not warnings_df.empty else pd.DataFrame({"warning": ["None"]}),
    )

    if plot_inputs:
        plot_df = pd.DataFrame({"plot_input_file": [str(p) for p in plot_inputs]})
        _add_dataframe_sheet(wb, "Report_files", plot_df)

    wb.save(outpath)
